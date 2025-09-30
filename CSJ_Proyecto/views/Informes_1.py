# Importar librerías principales
import streamlit as st  # Framework para crear dashboards web
import pandas as pd  # Manejo de DataFrames
import os  # Operaciones de sistema de archivos
import io  # Manejo de streams de datos (para Excel)
from openpyxl.utils.dataframe import dataframe_to_rows  # Convertir DataFrames a filas de Excel
from openpyxl.styles import Font  # Formato de celdas en Excel
from openpyxl import Workbook  # Crear libros de Excel
import re  # Expresiones regulares para limpiar HTML
from eventos_utils import obtener_archivos_por_anio_y_mes, meses_es  # Funciones personalizadas de eventos

# Función principal del dashboard
def main():
    # Configuración inicial de la página Streamlit
    st.set_page_config(page_title="Dashboard Clínico", layout="wide")
    st.title("📊 Dashboard Clínico - Análisis Descriptivo")

    # Carpeta donde se almacenan los archivos de datos
    CARPETA_DATOS = "data_subida"
    os.makedirs(CARPETA_DATOS, exist_ok=True)  # Crear carpeta si no existe

    # Diccionario inverso de meses (número -> nombre)
    meses_num_a_nombre = {v: k for k, v in meses_es.items()}

    # --- Sección de selección manual de año y mes ---
    with st.expander("📂 Selección manual de Año y Mes"):
        # Obtener archivos disponibles por año y mes
        archivos_por_anio_mes = obtener_archivos_por_anio_y_mes(CARPETA_DATOS)

        # Advertencia si no hay archivos
        if not archivos_por_anio_mes:
            st.warning("⚠️ No se encontraron archivos válidos con año y mes reconocibles.")
            st.stop()

        # Selección de año
        anios_disponibles = sorted(archivos_por_anio_mes.keys(), reverse=True)
        anio_default = st.session_state.get("anio_seleccionado", anios_disponibles[0])
        if anio_default not in anios_disponibles:
            anio_default = anios_disponibles[0]

        anio_seleccionado = st.selectbox(
            "Selecciona el año:",
            options=anios_disponibles,
            index=anios_disponibles.index(anio_default),
            key="anio_seleccionado"
        )

        # Selección de mes
        meses_disponibles = sorted(archivos_por_anio_mes[anio_seleccionado].keys())
        meses_nombres = [nombre for nombre, num in meses_es.items() if num in meses_disponibles]
        meses_nombres_ordenados = sorted(meses_nombres, key=lambda x: meses_es[x])

        mes_default_num = st.session_state.get("mes_seleccionado", meses_disponibles[0])
        nombre_mes_default = meses_num_a_nombre.get(mes_default_num, meses_nombres_ordenados[0])
        if nombre_mes_default not in meses_nombres_ordenados:
            nombre_mes_default = meses_nombres_ordenados[0]

        mes_seleccionado_nombre = st.selectbox(
            "Selecciona el mes:",
            options=meses_nombres_ordenados,
            index=meses_nombres_ordenados.index(nombre_mes_default),
            key="mes_seleccionado_nombre"
        )
        mes_seleccionado = meses_es[mes_seleccionado_nombre]

        # Guardar selección en session_state
        st.session_state["mes_seleccionado"] = mes_seleccionado
        st.session_state["archivo_seleccionado"] = archivos_por_anio_mes[anio_seleccionado][mes_seleccionado]

        # Mostrar selección
        st.write(f"Has seleccionado: Año {anio_seleccionado}, Mes {mes_seleccionado_nombre.capitalize()}")

    # Ruta del archivo seleccionado
    archivo_cargado = st.session_state.get("archivo_seleccionado", None)
    ruta_archivo = os.path.join(CARPETA_DATOS, archivo_cargado) if archivo_cargado else None

    # Validar existencia del archivo
    if not archivo_cargado or not os.path.exists(ruta_archivo):
        st.warning("⚠️ No se encontró archivo seleccionado o no existe en la carpeta `data_subida`.")
        st.stop()

    # Cargar archivo Excel en DataFrame
    df = pd.read_excel(ruta_archivo)
    st.success(f"✅ Archivo cargado automáticamente: {archivo_cargado}")
    st.markdown(f"**Total de registros:** {df.shape[0]}")
    st.markdown(f"**Total de columnas:** {df.shape[1]}")

    # --- Estilo HTML para tablas ---
    estilo_tablas = """
    <style>
    .table-custom {
        border-collapse: collapse;
        width: 100%;
        font-family: Arial, sans-serif;
        font-size: 14px;
    }
    .table-custom th {
        background-color: #cce5ff;
        padding: 8px;
        border: 1px solid black;
        text-align: center;
    }
    .table-custom td {
        padding: 6px;
        border: 1px solid black;
        text-align: center;
    }
    .table-custom tr:last-child,
    .table-custom tr:nth-last-child(2) {
        background-color: #ffe5cc;
        font-weight: bold;
    }
    </style>
    """
    st.markdown(estilo_tablas, unsafe_allow_html=True)

    # --- Función para filtrar eventos adversos ---
    def filtrar_adversos(df):
        return df[df["Clasificación"].str.strip().str.lower() == "adverso"]

    # --- Vista previa de datos ---
    st.subheader("🗂️ Vista previa de los datos")
    st.dataframe(df)

    # --- Tabla con columnas seleccionadas ---
    if {"Servicio Ocurrencia", "Evento", "Clasificación"}.issubset(df.columns):
        st.subheader("🔎 Tabla con columnas seleccionadas")

        clasificaciones = sorted(df["Clasificación"].dropna().unique())
        clasificacion_pred = "Adverso" if "Adverso" in clasificaciones else clasificaciones[0]

        clasificacion_sel = st.radio(
            "Filtrar por Clasificación:",
            options=clasificaciones,
            index=clasificaciones.index(clasificacion_pred),
            horizontal=True
        )

        df_filtrado = df[df["Clasificación"].str.strip().str.lower() == clasificacion_sel.lower()]
        tabla_final_clasif = df_filtrado[["Servicio Ocurrencia", "Evento", "Clasificación"]].copy()

        st.dataframe(tabla_final_clasif)

        # --- Resumen de Clasificación General ---
        if "Clasificación" in df.columns:
            st.subheader("📋 Tabla de Clasificación General")

            resumen_clasif = df["Clasificación"].value_counts().reset_index()
            resumen_clasif.columns = ["Clasificación", "Cantidad"]

            total = resumen_clasif["Cantidad"].sum()
            resumen_clasif["Peso,%"] = (resumen_clasif["Cantidad"] / total * 100).round(1).astype(str) + "%"

            fila_total = pd.DataFrame({
                "Clasificación": ["<b>Total general</b>"],
                "Cantidad": [f"<b>{total}</b>"],
                "Peso,%": ["<b>100.0%</b>"]
            })
            tabla_resumen_clasif = pd.concat([resumen_clasif, fila_total], ignore_index=True)

            html = '<table class="table-custom"><thead><tr>'
            for col in tabla_resumen_clasif.columns:
                html += f"<th>{col}</th>"
            html += "</tr></thead><tbody>"
            for _, row in tabla_resumen_clasif.iterrows():
                html += "<tr>"
                for col in tabla_resumen_clasif.columns:
                    html += f"<td>{row[col]}</td>"
                html += "</tr>"
            html += "</tbody></table>"

            st.markdown("""<style>
            .table-custom { border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 14px; }
            .table-custom th { background-color: #cce5ff; padding: 8px; border: 1px solid black; text-align: center; }
            .table-custom td { padding: 6px; border: 1px solid black; text-align: center; }
            .table-custom tr:last-child { background-color: #ffe5cc; font-weight: bold; }
            </style>""", unsafe_allow_html=True)

            st.markdown(html, unsafe_allow_html=True)

    # --- Pivot Table: Totales por Servicio ---
    if {"Servicio Ocurrencia", "Clasificación"}.issubset(df.columns):
        pivot = pd.pivot_table(
            df,
            index="Servicio Ocurrencia",
            columns="Clasificación",
            aggfunc="size",
            fill_value=0
        )
        pivot["Total general"] = pivot.sum(axis=1)
        total_global = pivot["Total general"].sum()
        pivot["Peso%"] = pivot["Total general"] / total_global * 100
        pivot["Peso%"] = pivot["Peso%"].map(lambda x: f"{x:.1f}%")
        totales_col = pivot.drop(columns=["Peso%"]).sum()
        totales_col["Peso%"] = "100,0%"
        peso_col = pivot.drop(columns=["Total general", "Peso%"]).sum()
        peso_col = (peso_col / total_global * 100).map(lambda x: f"{x:.1f}%")
        peso_col["Total general"] = "100,0%"
        peso_col["Peso%"] = ""
        pivot.loc["<b>Total general</b>"] = totales_col
        pivot.loc["<b>Peso,%</b>"] = peso_col
        tabla_servicio = pivot.reset_index()
        st.subheader("📋 Totales por Servicio")
        st.markdown(tabla_servicio.to_html(index=False, escape=False, classes="table-custom"), unsafe_allow_html=True)

    # --- Eventos por Servicio con filtros ---
    if {"Evento", "Servicio Ocurrencia", "Clasificación"}.issubset(df.columns):
        df_adversos = filtrar_adversos(df)

        st.subheader("📋 Tabla de Eventos por Servicio")
        
        # Filtro múltiple: Servicio Ocurrencia
        servicios_disponibles = df_adversos["Servicio Ocurrencia"].dropna().unique()
        servicios_seleccionados = st.multiselect("🏥 Selecciona Servicios", sorted(servicios_disponibles))
        if servicios_seleccionados:
            df_adversos = df_adversos[df_adversos["Servicio Ocurrencia"].isin(servicios_seleccionados)]

        # Filtro múltiple: Evento
        eventos_disponibles = df_adversos["Evento"].dropna().unique()
        eventos_seleccionados = st.multiselect("🎯 Selecciona Eventos", sorted(eventos_disponibles))
        if eventos_seleccionados:
            df_adversos = df_adversos[df_adversos["Evento"].isin(eventos_seleccionados)]

        # Pivot Table de Eventos
        pivot_evento = pd.pivot_table(
            df_adversos,
            index="Evento",
            columns="Servicio Ocurrencia",
            aggfunc="size",
            fill_value=0
        )

        # Eliminar filas y columnas vacías
        pivot_evento = pivot_evento.loc[:, (pivot_evento != 0).any(axis=0)]
        pivot_evento = pivot_evento[(pivot_evento.T != 0).any()]

        # Si hay datos, agregar totales y peso
        if not pivot_evento.empty:
            pivot_evento["Total"] = pivot_evento.sum(axis=1)
            total_general = pivot_evento["Total"].sum()
            pivot_evento["Peso%"] = (pivot_evento["Total"] / total_general * 100).map(lambda x: f"{x:.1f}%")

            columnas_evento = pivot_evento.columns.drop(["Total", "Peso%"])
            totales_col = pivot_evento[columnas_evento].sum()
            totales_col["Total"] = total_general
            totales_col["Peso%"] = "100.0%"

            peso_col = (pivot_evento[columnas_evento].sum() / total_general * 100).map(lambda x: f"{x:.1f}%")
            peso_col["Total"] = "100.0%"
            peso_col["Peso%"] = ""

            fila_total = pd.DataFrame([totales_col], index=["<b>Total general</b>"])
            fila_peso = pd.DataFrame([peso_col], index=["<b>Peso,%</b>"])
            tabla_eventos = pd.concat([pivot_evento, fila_total, fila_peso])
            tabla_eventos = tabla_eventos.reset_index().rename(columns={"index": "Evento"})

            # Construir tabla HTML
            html = '<table class="table-custom"><thead><tr>'
            for col in tabla_eventos.columns:
                html += f"<th>{col}</th>"
            html += "</tr></thead><tbody>"
            for _, row in tabla_eventos.iterrows():
                html += "<tr>"
                for col in tabla_eventos.columns:
                    valor = row[col]
                    html += f'<td>{valor}</td>'
                html += "</tr>"
            html += "</tbody></table>"

            # CSS de tabla
            tabla_css = """
            <style>
            .table-custom {
                border-collapse: collapse;
                width: 100%;
                font-family: Arial, sans-serif;
                font-size: 14px;
            }
            .table-custom th {
                background-color: #d9e0f5;
                padding: 8px;
                border: 1px solid black;
                text-align: center;
            }
            .table-custom td {
                padding: 6px;
                border: 1px solid black;
                text-align: center;
            }
            .table-custom tr:nth-last-child(2),
            .table-custom tr:last-child {
                background-color: #ffe5cc !important;
                font-weight: bold;
            }
            </style>
            """
            st.markdown(tabla_css, unsafe_allow_html=True)
            st.markdown(html, unsafe_allow_html=True)
        else:
            st.info("⚠️ No hay datos que coincidan con los filtros.")

    # --- Resumen de eventos adversos principales ---
    if {"Servicio Ocurrencia", "Evento", "Clasificación"}.issubset(df.columns): 
        df_adversos = df[df["Clasificación"].str.strip().str.lower() == "adverso"]
        servicio_top = df_adversos["Servicio Ocurrencia"].value_counts().idxmax()
        cantidad_servicio_top = df_adversos["Servicio Ocurrencia"].value_counts().max()
        evento_top = df_adversos[df_adversos["Servicio Ocurrencia"] == servicio_top]["Evento"].value_counts().idxmax()
        cantidad_evento_top = df_adversos[df_adversos["Servicio Ocurrencia"] == servicio_top]["Evento"].value_counts().max()

        resumen_columnas = pd.DataFrame({
            "Servicio con más eventos adversos": [servicio_top],
            "Cantidad total de eventos": [cantidad_servicio_top],
            "Evento más frecuente en ese servicio": [evento_top],
            "Frecuencia total de eventos": [cantidad_evento_top]
        })

        st.subheader("📌 Resumen de Eventos Adversos Principales")
        st.markdown("""<style>
        .resumen-evento-col { border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 15px; margin-bottom: 15px; table-layout: fixed; }
        .resumen-evento-col th { background-color: #e6f2ff; padding: 8px; border: 1px solid black; text-align: center; }
        .resumen-evento-col td { padding: 6px 10px; border: 1px solid black; text-align: center; word-wrap: break-word; }
        </style>""", unsafe_allow_html=True)
        st.markdown(resumen_columnas.to_html(index=False, escape=False, classes="resumen-evento-col"), unsafe_allow_html=True)

    # --- Tabla de eventos adversos por frecuencia ---
    if {"Evento", "Clasificación"}.issubset(df.columns):
        df_adversos = filtrar_adversos(df)
        conteo_eventos = df_adversos["Evento"].value_counts().reset_index()
        conteo_eventos.columns = ["Evento", "N° Evento Adverso"]
        total_eventos = conteo_eventos["N° Evento Adverso"].sum()
        conteo_eventos["Peso,%"] = conteo_eventos["N° Evento Adverso"] / total_eventos * 100
        conteo_eventos["Peso,%"] = conteo_eventos["Peso,%"].map(lambda x: f"{x:.1f}%")
        fila_total = pd.DataFrame({
            "Evento": ["<b>Total general</b>"],
            "N° Evento Adverso": [f"<b>{total_eventos}</b>"],
            "Peso,%": ["<b>100.0%</b>"]
        })
        tabla_eventos_adversos = pd.concat([conteo_eventos, fila_total], ignore_index=True)
        st.subheader("📋 Tabla de Evento Adverso por Frecuencia")
        st.markdown("""<style>
        .tabla-eventos-adversos { border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 14px; }
        .tabla-eventos-adversos th { background-color: #cce5ff; padding: 8px; border: 1px solid black; text-align: center; }
        .tabla-eventos-adversos td { padding: 6px; border: 1px solid black; text-align: center; }
        .tabla-eventos-adversos tr:last-child { background-color: #ffe5cc; font-weight: bold; }
        </style>""", unsafe_allow_html=True)
        st.markdown(tabla_eventos_adversos.to_html(index=False, escape=False, classes="tabla-eventos-adversos"), unsafe_allow_html=True)

    # --- Pivot de clasificación por Servicio ---
    if {"Servicio Ocurrencia", "Clasificación"}.issubset(df.columns):
        pivot_clasif = pd.pivot_table(
            df,
            index="Servicio Ocurrencia",
            columns="Clasificación",
            aggfunc="size",
            fill_value=0
        )
        pivot_clasif["Total"] = pivot_clasif.sum(axis=1)
        total_gral = pivot_clasif["Total"].sum()
        pivot_clasif["Peso%"] = pivot_clasif["Total"] / total_gral * 100
        pivot_clasif["Peso%"] = pivot_clasif["Peso%"].map(lambda x: f"{x:.1f}%")
        totales_columna = pivot_clasif.drop(columns=["Peso%"]).sum()
        totales_columna["Peso%"] = "100.0%"
        peso_columna = (pivot_clasif.drop(columns=["Total", "Peso%"]).sum() / total_gral * 100).map(lambda x: f"{x:.1f}%")
        peso_columna["Total"] = "100.0%"
        peso_columna["Peso%"] = ""
        pivot_clasif.loc["<b>Total general</b>"] = totales_columna
        pivot_clasif.loc["<b>Peso,%</b>"] = peso_columna
        tabla_cuenta_clasificacion = pivot_clasif.reset_index()
        st.subheader("📋 Clasificación por Servicio")
        st.markdown(tabla_cuenta_clasificacion.to_html(index=False, escape=False, classes="table-custom"), unsafe_allow_html=True)

    # --- Botón de navegación ---
#    if st.button("📈 Ir a Gráficos"):
#        st.session_state["pending_nav_change"] = True
#        st.session_state["menu_option"] = "📊 Gráficos"
#        st.rerun()

    # --- Función para limpiar HTML de las tablas ---
    def limpiar_html(df):
        return df.applymap(lambda x: re.sub(r'<.*?>', '', str(x)) if isinstance(x, str) else x)

    st.subheader("📥 Exportar Tablas del Dashboard")

    # --- Preparar Excel ---
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tablas Clínicas"
    fila = 1

    # Función para agregar tabla al Excel
    def agregar_tabla(nombre, tabla, fila_inicial):
        tabla = limpiar_html(tabla)
        ws.cell(row=fila_inicial, column=1, value=nombre).font = Font(bold=True)
        fila_actual = fila_inicial + 1
        for r in dataframe_to_rows(tabla, index=False, header=True):
            for c_idx, val in enumerate(r, 1):
                ws.cell(row=fila_actual, column=c_idx, value=val)
            fila_actual += 1
        fila_actual += 1
        return fila_actual

    # Diccionario de tablas a exportar
    tablas = {
        "Resumen Clasificación": "tabla_final_clasif",
        "Totales por Servicio": "tabla_servicio",
        "Eventos por Servicio": "tabla_eventos",
        "Eventos Adversos Frecuencia": "tabla_eventos_adversos",
        "Resumen Principal": "resumen_columnas",
        "Clasificación por Servicio": "tabla_cuenta_clasificacion"
    }

    # Agregar todas las tablas al Excel
    for nombre, var in tablas.items():
        if var in locals():
            tabla = locals()[var]
            fila = agregar_tabla(nombre, tabla, fila)

    # Guardar y preparar botón de descarga
    wb.save(output)
    st.download_button(
        label="📤 Descargar Excel",
        data=output.getvalue(),
        file_name="dashboard_clinico.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Ejecutar la función principal
if __name__ == "__main__":
    main()
